print("test4開始")

from openpyxl import load_workbook

input_file = "sample.xlsx"
output_file = "完成_デキスパート入力.xlsx"

values = [
    1.418, 1.400, 1.409, 1.409, 1.415,
    1.419, 1.421, 1.412, 1.403, 1.403,
    1.416, 1.406, 1.409, 1.403, 1.406,
    1.406, 1.404, 1.410, 1.415, 1.406,
    1.411, 1.406, 1.400, 1.415, 1.420,
    1.411, 1.419, 1.416, 1.406,
]

wb = load_workbook(input_file)
print("Excelを開きました")

ws = wb["測定結果表"]

# まずは D列に入れてみる
start_row = 9
target_col = "D"

for i, value in enumerate(values):
    ws[f"{target_col}{start_row + i}"] = value

wb.save(output_file)

print("保存しました:", output_file)