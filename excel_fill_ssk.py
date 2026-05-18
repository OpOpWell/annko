import os
import csv
import glob
from openpyxl import load_workbook

template_excel = "測定結果表.xlsx"
csv_folder = "ssk_output"
output_folder = "excel_output"

os.makedirs(output_folder, exist_ok=True)


def load_meta():
    meta_path = os.path.join(csv_folder, "測定情報.csv")
    meta = {}

    if not os.path.exists(meta_path):
        return meta

    with open(meta_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for row in reader:
            taban = row.get("田番", "")
            meta[taban] = row

    return meta


def find_no_rows(ws):
    no_rows = {}

    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value

        if value is None:
            continue

        text = str(value).strip()

        if text.startswith("No."):
            try:
                no = int(text.replace("No.", ""))
                no_rows[no] = row
            except:
                pass

        elif text.isdigit():
            no_rows[int(text)] = row

    return no_rows


def write_label_value(ws, label, value):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value

            if cell_value is None:
                continue

            if label in str(cell_value):
                ws.cell(row=row, column=col + 1).value = value
                return


meta_by_taban = load_meta()

csv_files = glob.glob(os.path.join(csv_folder, "田番*.csv"))

for csv_file in csv_files:

    taban = os.path.basename(csv_file)
    taban = taban.replace("田番", "").replace(".csv", "")

    print(f"処理中: 田番{taban}")

    wb = load_workbook(template_excel)
    ws = wb.active

    meta = meta_by_taban.get(taban, {})

    design_value = meta.get("平均値", "")
    spec_value = meta.get("規格値", "")
    target_value = meta.get("社内規格値", "")

    if design_value == "":
        design_value = meta.get("average", "")

    # 規格値などを書ける場合
    write_label_value(ws, "規格値", spec_value)
    write_label_value(ws, "社内規格値", target_value)

    no_rows = find_no_rows(ws)

    with open(csv_file, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for row in reader:
            point = row["測点"]
            actual_value = row["実測値"]

            if actual_value == "":
                continue

            no = int(point.replace("No.", ""))

            if no not in no_rows:
                continue

            excel_row = no_rows[no]

            # B列 = 設計値
            ws.cell(row=excel_row, column=2).value = float(design_value)

            # C列 = 実測値
            ws.cell(row=excel_row, column=3).value = float(actual_value)

            # D列 = 差 mm
            ws.cell(row=excel_row, column=4).value = (
                f"=ROUND((B{excel_row}-C{excel_row})*1000,0)"
            )

            ws.cell(row=excel_row, column=2).number_format = "0.000"
            ws.cell(row=excel_row, column=3).number_format = "0.000"
            ws.cell(row=excel_row, column=4).number_format = "0"

    output_excel = os.path.join(
        output_folder,
        f"完成_田番{taban}.xlsx"
    )

    wb.save(output_excel)

    print("保存:", output_excel)

print("全部完了")