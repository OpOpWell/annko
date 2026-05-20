import tkinter as tk
from tkinter import messagebox, scrolledtext
from openai import OpenAI
from dotenv import load_dotenv

import base64
import os
import csv
import re
import json
import shutil
import threading

from openpyxl import Workbook
from openpyxl.styles import PatternFill

# =========================
# .env 読込
# =========================

load_dotenv(dotenv_path=".env")

client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY")
)

# =========================
# 基本フォルダ
# =========================

BASE_FOLDER = r"C:\Users\user\foolder\634"

IMAGES_ALL = os.path.join(BASE_FOLDER, "images_all")

PHOTO_SORTED = os.path.join(BASE_FOLDER, "photo_sorted")

SELECTED_PHOTOS = os.path.join(BASE_FOLDER, "selected_photos")

SELECTED_DEKIGATA = os.path.join(SELECTED_PHOTOS, "出来形使用写真")
SELECTED_WORK = os.path.join(SELECTED_PHOTOS, "作業使用写真")
SELECTED_MACHINE = os.path.join(SELECTED_PHOTOS, "重機使用写真")
SELECTED_BOARD = os.path.join(SELECTED_PHOTOS, "黒板使用写真")
SELECTED_COMPLETE = os.path.join(SELECTED_PHOTOS, "完成使用写真")
SELECTED_BEFORE = os.path.join(SELECTED_PHOTOS, "着工前使用写真")
SELECTED_SUBMIT = os.path.join(SELECTED_PHOTOS, "提出候補")

CSV_OUTPUT = os.path.join(BASE_FOLDER, "csv_output")
CSV_OK = os.path.join(CSV_OUTPUT, "ok")
CSV_WARNING = os.path.join(CSV_OUTPUT, "warning")
CSV_ERROR = os.path.join(CSV_OUTPUT, "error")

REPORTS = os.path.join(BASE_FOLDER, "reports")

WORK = os.path.join(BASE_FOLDER, "work")
ERROR_FOLDER = os.path.join(WORK, "error")

PHOTO_CATEGORIES = [
    "出来形測定写真",
    "作業写真",
    "重機写真",
    "黒板写真",
    "完成写真",
    "着工前",
    "除外"
]

# =========================
# フォルダ作成
# =========================

def make_folders():

    folders = [
        IMAGES_ALL,
        PHOTO_SORTED,
        SELECTED_PHOTOS,
        SELECTED_DEKIGATA,
        SELECTED_WORK,
        SELECTED_MACHINE,
        SELECTED_BOARD,
        SELECTED_COMPLETE,
        SELECTED_BEFORE,
        SELECTED_SUBMIT,
        CSV_OUTPUT,
        CSV_OK,
        CSV_WARNING,
        CSV_ERROR,
        REPORTS,
        WORK,
        ERROR_FOLDER
    ]

    for folder in folders:
        os.makedirs(folder, exist_ok=True)

    for category in PHOTO_CATEGORIES:
        os.makedirs(
            os.path.join(PHOTO_SORTED, category),
            exist_ok=True
        )

make_folders()

# =========================
# マスタ読込
# =========================

expected_counts = {}

with open("master_taban.csv", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)

    for row in reader:
        expected_counts[str(row["田番"]).strip()] = int(row["測点数"])

allowed_items = []

with open("master_items.csv", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)

    for row in reader:
        allowed_items.append(row["測定項目"].strip())

work_master = {}

with open("master_work.csv", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)

    for row in reader:
        item = row["測定項目"].strip()
        work_master[item] = row["工種"].strip()

# =========================
# 色
# =========================

GREEN_FILL = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

YELLOW_FILL = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid"
)

RED_FILL = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

# =========================
# 共通関数
# =========================

def safe_copy(src, dst_folder):

    os.makedirs(dst_folder, exist_ok=True)

    filename = os.path.basename(src)

    dst = os.path.join(dst_folder, filename)

    shutil.copy2(src, dst)

    return dst


def image_to_base64(image_path):

    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def extract_json(text):

    text = text.replace("```json", "")
    text = text.replace("```", "")

    match = re.search(r"\{.*\}", text, re.DOTALL)

    if not match:
        return {}

    try:
        return json.loads(match.group())
    except:
        return {}


def to_mm_text(value):

    try:
        return str(int(round(float(value) * 1000)))
    except:
        return ""

# =========================
# AI写真分類
# =========================

def classify_photo(image_path):

    try:
        base64_image = image_to_base64(image_path)

        response = client.responses.create(
            model="gpt-4.1-mini",
            temperature=0,
            input=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": """
この工事写真を分類してください。

返却はJSONのみ。
説明文は禁止。

分類は必ず次から1つ選んでください。

- 出来形測定写真
- 作業写真
- 重機写真
- 黒板写真
- 完成写真
- 着工前
- 除外

判断基準:
- 出来形測定写真: 測定表、手書き数値、測点番号、田番、出来形管理の写真
- 作業写真: 人が作業している、施工状況が分かる写真
- 重機写真: バックホウ、ローラー、ダンプなど重機が主役
- 黒板写真: 工事黒板が主
- 完成写真: 施工後、完成後の状態
- 着工前: 施工前の状態
- 除外: ピンぼけ、関係ない、判別不能、暗すぎる、地面だけ

返却形式:
{
  "category": "出来形測定写真",
  "reason": "測定表と手書き数値が写っているため"
}
"""
                        },
                        {
                            "type": "input_image",
                            "image_url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    ]
                }
            ]
        )

        data = extract_json(response.output_text.strip())

        category = str(data.get("category", "除外")).strip()
        reason = str(data.get("reason", "")).strip()

        if category not in PHOTO_CATEGORIES:
            category = "除外"

        return category, reason

    except Exception as e:
        return "除外", f"分類エラー: {e}"

# =========================
# SSK OCR
# =========================

def gpt_ocr(image_path):

    base64_image = image_to_base64(image_path)

    response = client.responses.create(
        model="gpt-4.1-mini",
        temperature=0,
        input=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": """
このSSK出来形測定写真から情報を読み取ってください。

重要:
- 田番は上部の「田番」欄の数字
- 大きく丸で書かれた手書き数字は田番にしない
- 番号と実測値を必ず対応させる
- No.1を飛ばさない
- 6と8を間違えやすいので注意する

測定項目は次の候補から選択してください。

- 均平度
- 幅
- 厚さ
- 基準高
- 法長
- 延長
- 深さ
- 天端高
- 出来高
- 掘削深
- 盛土高

読む項目:
- 田番
- 測定項目
- 平均値
- 番号と手書き実測値

条件:
- 空欄は無視
- 実測値は小数3桁
- JSONのみ返す
- 説明不要
- 推測しない

返却形式:
{
  "taban": "9",
  "measurement_item": "均平度",
  "average": "1.452",
  "values": [
    {"no": 1, "value": 1.446},
    {"no": 2, "value": 1.456}
  ]
}
"""
                    },
                    {
                        "type": "input_image",
                        "image_url": f"data:image/jpeg;base64,{base64_image}"
                    }
                ]
            }
        ]
    )

    raw_text = response.output_text.strip()

    return raw_text, extract_json(raw_text)

# =========================
# 自動検査
# =========================

def validate_result(taban, measurement_item, values):

    status = "OK"
    reasons = []

    if taban not in expected_counts:
        return "ERROR", ["存在しない田番"]

    if measurement_item not in allowed_items:
        status = "ERROR"
        reasons.append("測定項目不明")

    expected = expected_counts[taban]

    nos = []

    for item in values:
        try:
            nos.append(int(item["no"]))
        except:
            pass

    nos = sorted(nos)

    if len(nos) != expected:
        status = "WARNING"
        reasons.append(f"測点数不一致 {len(nos)}/{expected}")

    expected_nos = list(range(1, expected + 1))

    if nos != expected_nos:
        status = "ERROR"
        reasons.append("測点連番異常")

    for item in values:
        try:
            value = float(item["value"])

            if value < 1.300 or value > 1.600:
                status = "ERROR"
                reasons.append(f"範囲外値 {value}")

        except:
            status = "ERROR"
            reasons.append("数値変換エラー")

    return status, reasons

# =========================
# Excelレポート
# =========================

def create_report_excel(rows):

    wb = Workbook()
    ws = wb.active
    ws.title = "OCRチェック"

    headers = [
        "画像名",
        "分類",
        "田番",
        "測定項目",
        "状態",
        "理由",
        "保存先"
    ]

    ws.append(headers)

    for row in rows:
        ws.append(row)

        status = row[4]

        current_row = ws.max_row

        if status == "OK":
            fill = GREEN_FILL
        elif status == "WARNING":
            fill = YELLOW_FILL
        else:
            fill = RED_FILL

        for cell in ws[current_row]:
            cell.fill = fill

    save_path = os.path.join(REPORTS, "OCRチェック.xlsx")

    wb.save(save_path)

    return save_path

# =========================
# CSV保存
# =========================

def save_taban_csv(save_folder, taban, rows):

    os.makedirs(save_folder, exist_ok=True)

    csv_path = os.path.join(save_folder, f"田番{taban}.csv")

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)

        writer.writerow([
            "測点",
            "工種名",
            "測定項目",
            "設計値",
            "実測値1"
        ])

        writer.writerows(rows)

    return csv_path

# =========================
# 使用写真コピー
# =========================

def copy_selected_by_category(image_path, category, options):

    copied_paths = []

    if category == "出来形測定写真" and options["copy_dekigata"]:
        copied_paths.append(
            safe_copy(image_path, SELECTED_DEKIGATA)
        )

    if category == "作業写真" and options["copy_work"]:
        copied_paths.append(
            safe_copy(image_path, SELECTED_WORK)
        )

    if category == "重機写真" and options["copy_machine"]:
        copied_paths.append(
            safe_copy(image_path, SELECTED_MACHINE)
        )

    if category == "黒板写真" and options["copy_board"]:
        copied_paths.append(
            safe_copy(image_path, SELECTED_BOARD)
        )

    if category == "完成写真" and options["copy_complete"]:
        copied_paths.append(
            safe_copy(image_path, SELECTED_COMPLETE)
        )

    if category == "着工前" and options["copy_before"]:
        copied_paths.append(
            safe_copy(image_path, SELECTED_BEFORE)
        )

    if options["make_submit"]:
        if category in [
            "出来形測定写真",
            "作業写真",
            "重機写真",
            "黒板写真",
            "完成写真",
            "着工前"
        ]:
            copied_paths.append(
                safe_copy(image_path, SELECTED_SUBMIT)
            )

    return copied_paths

# =========================
# メイン処理
# =========================

def run_all(log, options):

    make_folders()

    report_rows = []

    image_files = sorted([
        f for f in os.listdir(IMAGES_ALL)
        if f.lower().endswith((".jpg", ".jpeg", ".png"))
    ])

    if not image_files:
        log("画像がありません")
        return

    log("AI写真分類 + SSK OCR 開始")
    log(f"対象枚数: {len(image_files)}")

    for index, image_name in enumerate(image_files, start=1):

        image_path = os.path.join(IMAGES_ALL, image_name)

        log(f"\n[{index}/{len(image_files)}] 処理中: {image_name}")

        category, reason = classify_photo(image_path)

        log(f"分類: {category}")
        log(f"理由: {reason}")

        sorted_path = safe_copy(
            image_path,
            os.path.join(PHOTO_SORTED, category)
        )

        selected_paths = copy_selected_by_category(
            image_path,
            category,
            options
        )

        for path in selected_paths:
            log(f"使用写真コピー: {path}")

        if category != "出来形測定写真":
            report_rows.append([
                image_name,
                category,
                "",
                "",
                "OK",
                reason,
                sorted_path
            ])
            continue

        if not options["ocr_dekigata"]:
            report_rows.append([
                image_name,
                category,
                "",
                "",
                "OK",
                "OCR未実行",
                sorted_path
            ])
            continue

        log("OCR開始")

        try:
            raw_text, result = gpt_ocr(image_path)

            log("GPT OCR結果")
            log(raw_text)

        except Exception as e:
            log(f"OCRエラー: {e}")

            safe_copy(image_path, ERROR_FOLDER)

            report_rows.append([
                image_name,
                category,
                "",
                "",
                "ERROR",
                f"OCRエラー: {e}",
                sorted_path
            ])

            continue

        taban = str(result.get("taban", "")).strip()
        measurement_item = str(result.get("measurement_item", "不明")).strip()
        values = result.get("values", [])

        status, reasons = validate_result(
            taban,
            measurement_item,
            values
        )

        reason_text = " / ".join(reasons)

        log(f"判定: {status}")
        log(f"理由: {reason_text}")

        if status == "OK":
            csv_folder = CSV_OK
        elif status == "WARNING":
            csv_folder = CSV_WARNING
        else:
            csv_folder = CSV_ERROR

        value_dict = {}

        for item in values:
            try:
                no = int(item["no"])
                value = float(item["value"])

                value_dict[no] = value

            except:
                pass

        expected = expected_counts.get(taban, 0)

        design_value = to_mm_text(result.get("average", ""))

        work_type = work_master.get(
            measurement_item,
            "整地工"
        )

        rows = []

        for no in range(1, expected + 1):

            if no in value_dict:
                value_text = to_mm_text(value_dict[no])
            else:
                value_text = ""

            rows.append([
                f"No.{no}",
                f"{work_type}田番{taban}",
                measurement_item,
                design_value,
                value_text
            ])

        if taban:
            csv_path = save_taban_csv(
                csv_folder,
                taban,
                rows
            )

            log(f"CSV保存: {csv_path}")
        else:
            csv_path = ""

        report_rows.append([
            image_name,
            category,
            taban,
            measurement_item,
            status,
            reason_text,
            csv_path
        ])

    excel_path = create_report_excel(report_rows)

    log("\n全部完了")
    log(f"Excelレポート: {excel_path}")

    messagebox.showinfo(
        "完了",
        "写真分類・使用写真コピー・OCR・CSV作成が完了しました"
    )

# =========================
# GUI
# =========================

root = tk.Tk()
root.title("杏子 all in one")
root.geometry("980x820")

ocr_dekigata_var = tk.BooleanVar(value=True)
copy_dekigata_var = tk.BooleanVar(value=True)
copy_work_var = tk.BooleanVar(value=True)
copy_machine_var = tk.BooleanVar(value=True)
copy_board_var = tk.BooleanVar(value=True)
copy_complete_var = tk.BooleanVar(value=True)
copy_before_var = tk.BooleanVar(value=True)
make_submit_var = tk.BooleanVar(value=True)

def write_log(message):
    log_text.insert(tk.END, str(message) + "\n")
    log_text.see(tk.END)

def start():

    options = {
        "ocr_dekigata": ocr_dekigata_var.get(),
        "copy_dekigata": copy_dekigata_var.get(),
        "copy_work": copy_work_var.get(),
        "copy_machine": copy_machine_var.get(),
        "copy_board": copy_board_var.get(),
        "copy_complete": copy_complete_var.get(),
        "copy_before": copy_before_var.get(),
        "make_submit": make_submit_var.get()
    }

    thread = threading.Thread(
        target=run_all,
        args=(write_log, options)
    )

    thread.start()

tk.Label(
    root,
    text="杏子 all in one\n写真分類 → 使用写真コピー → 出来形OCR → デキスパートCSV",
    font=("Meiryo", 14)
).pack(pady=10)

tk.Label(
    root,
    text=f"元写真フォルダ:\n{IMAGES_ALL}"
).pack(pady=5)

frame = tk.LabelFrame(
    root,
    text="実行オプション",
    padx=10,
    pady=10
)

frame.pack(pady=10)

tk.Checkbutton(
    frame,
    text="出来形測定写真をOCRする",
    variable=ocr_dekigata_var
).grid(row=0, column=0, sticky="w")

tk.Checkbutton(
    frame,
    text="出来形測定写真を使用写真へコピー",
    variable=copy_dekigata_var
).grid(row=1, column=0, sticky="w")

tk.Checkbutton(
    frame,
    text="作業写真を使用写真へコピー",
    variable=copy_work_var
).grid(row=2, column=0, sticky="w")

tk.Checkbutton(
    frame,
    text="重機写真を使用写真へコピー",
    variable=copy_machine_var
).grid(row=3, column=0, sticky="w")

tk.Checkbutton(
    frame,
    text="黒板写真を使用写真へコピー",
    variable=copy_board_var
).grid(row=4, column=0, sticky="w")

tk.Checkbutton(
    frame,
    text="完成写真を使用写真へコピー",
    variable=copy_complete_var
).grid(row=5, column=0, sticky="w")

tk.Checkbutton(
    frame,
    text="着工前写真を使用写真へコピー",
    variable=copy_before_var
).grid(row=6, column=0, sticky="w")

tk.Checkbutton(
    frame,
    text="提出候補フォルダにもコピー",
    variable=make_submit_var
).grid(row=7, column=0, sticky="w")

tk.Button(
    root,
    text="AI写真分類 + OCR + 使用写真抽出 開始",
    command=start,
    bg="blue",
    fg="white",
    height=2,
    width=40
).pack(pady=10)

log_text = scrolledtext.ScrolledText(
    root,
    width=130,
    height=35
)

log_text.pack(padx=10, pady=10)

root.mainloop()