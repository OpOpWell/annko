import tkinter as tk
from tkinter import messagebox, scrolledtext, filedialog
from openai import OpenAI
from dotenv import load_dotenv

import base64
import os
import csv
import re
import json
import shutil
import threading

from openpyxl import Workbook
from openpyxl.styles import PatternFill

# =========================
# .env 読込
# =========================

from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
load_dotenv(dotenv_path=SCRIPT_DIR / ".env")
load_dotenv(dotenv_path=".env")

client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY")
)

# =========================
# 基本フォルダ
# =========================

BASE_FOLDER = r"C:\Users\user\foolder\634"

IMAGES_ALL = os.path.join(BASE_FOLDER, "images_all")

PHOTO_SORTED = os.path.join(BASE_FOLDER, "photo_sorted")

SELECTED_PHOTOS = os.path.join(BASE_FOLDER, "selected_photos")

SELECTED_DEKIGATA = os.path.join(SELECTED_PHOTOS, "出来形使用写真")
SELECTED_WORK = os.path.join(SELECTED_PHOTOS, "作業使用写真")
SELECTED_MACHINE = os.path.join(SELECTED_PHOTOS, "重機使用写真")
SELECTED_BOARD = os.path.join(SELECTED_PHOTOS, "黒板使用写真")
SELECTED_COMPLETE = os.path.join(SELECTED_PHOTOS, "完成使用写真")
SELECTED_BEFORE = os.path.join(SELECTED_PHOTOS, "着工前使用写真")
SELECTED_SUBMIT = os.path.join(SELECTED_PHOTOS, "提出候補")

# =========================
# デキスパート写真ツリー用フォルダ
# =========================

DEKISPART_TREE = os.path.join(SELECTED_PHOTOS, "デキスパート取込")
TREE_BEFORE_COMPLETE = os.path.join(DEKISPART_TREE, "着工前及び完成写真")
TREE_WORK_STATUS = os.path.join(DEKISPART_TREE, "施工状況写真")
TREE_SAFETY = os.path.join(DEKISPART_TREE, "安全管理写真")
TREE_MATERIAL = os.path.join(DEKISPART_TREE, "使用材料写真")
TREE_QUALITY = os.path.join(DEKISPART_TREE, "品質管理写真")
TREE_DEKIGATA = os.path.join(DEKISPART_TREE, "出来形管理写真")
TREE_DISASTER = os.path.join(DEKISPART_TREE, "災害写真")
TREE_ACCIDENT = os.path.join(DEKISPART_TREE, "事故写真")
TREE_OTHER = os.path.join(DEKISPART_TREE, "その他")

# =========================
# 現場別デキスパート写真ツリーマスタ
# =========================
# master_tree.csv があれば、その中身を優先します。
# 列名: 写真区分,工種,種別,細別
# 例: 施工状況写真,整地工,整地工,整地仕上げ
# 例: 出来形管理写真,整地工,整地工,整地仕上げ

DEFAULT_TREE_ROWS = [
    {"写真区分": "施工状況写真", "工種": "整地工", "種別": "整地工", "細別": "整地仕上げ"},
    {"写真区分": "出来形管理写真", "工種": "整地工", "種別": "整地工", "細別": "整地仕上げ"},
    {"写真区分": "着工前及び完成写真", "工種": "", "種別": "", "細別": "着工前"},
    {"写真区分": "着工前及び完成写真", "工種": "", "種別": "", "細別": "完成"},
]

def load_tree_master():
    path = os.path.join(BASE_FOLDER, "master_tree.csv")

    if not os.path.exists(path):
        return DEFAULT_TREE_ROWS

    rows = []
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append({
                "写真区分": str(row.get("写真区分", "")).strip(),
                "工種": str(row.get("工種", "")).strip(),
                "種別": str(row.get("種別", "")).strip(),
                "細別": str(row.get("細別", "")).strip(),
            })

    return rows if rows else DEFAULT_TREE_ROWS

TREE_MASTER = load_tree_master()

def unique_values(key):
    values = []
    for row in TREE_MASTER:
        v = str(row.get(key, "")).strip()
        if v and v not in values:
            values.append(v)
    return values

WORK_TYPE_CANDIDATES = unique_values("工種") + ["未分類工種"]

CSV_OUTPUT = os.path.join(BASE_FOLDER, "csv_output")
CSV_OK = os.path.join(CSV_OUTPUT, "ok")
CSV_WARNING = os.path.join(CSV_OUTPUT, "warning")
CSV_ERROR = os.path.join(CSV_OUTPUT, "error")

REPORTS = os.path.join(BASE_FOLDER, "reports")

WORK = os.path.join(BASE_FOLDER, "work")
ERROR_FOLDER = os.path.join(WORK, "error")

PHOTO_CATEGORIES = [
    "出来形測定写真",
    "作業写真",
    "重機写真",
    "黒板写真",
    "完成写真",
    "着工前",
    "除外"
]

# =========================
# フォルダ作成
# =========================

def make_folders():

    folders = [
        IMAGES_ALL,
        PHOTO_SORTED,
        SELECTED_PHOTOS,
        SELECTED_DEKIGATA,
        SELECTED_WORK,
        SELECTED_MACHINE,
        SELECTED_BOARD,
        SELECTED_COMPLETE,
        SELECTED_BEFORE,
        SELECTED_SUBMIT,
        DEKISPART_TREE,
        TREE_BEFORE_COMPLETE,
        TREE_WORK_STATUS,
        TREE_SAFETY,
        TREE_MATERIAL,
        TREE_QUALITY,
        TREE_DEKIGATA,
        TREE_DISASTER,
        TREE_ACCIDENT,
        TREE_OTHER,
        CSV_OUTPUT,
        CSV_OK,
        CSV_WARNING,
        CSV_ERROR,
        REPORTS,
        WORK,
        ERROR_FOLDER
    ]

    for folder in folders:
        os.makedirs(folder, exist_ok=True)

    for category in PHOTO_CATEGORIES:
        os.makedirs(
            os.path.join(PHOTO_SORTED, category),
            exist_ok=True
        )

make_folders()

# =========================
# マスタ読込
# =========================

expected_counts = {}

with open("master_taban.csv", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)

    for row in reader:
        expected_counts[str(row["田番"]).strip()] = int(row["測点数"])

allowed_items = []

with open("master_items.csv", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)

    for row in reader:
        allowed_items.append(row["測定項目"].strip())

work_master = {}

with open("master_work.csv", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)

    for row in reader:
        item = row["測定項目"].strip()
        work_master[item] = row["工種"].strip()

# =========================
# 色
# =========================

GREEN_FILL = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

YELLOW_FILL = PatternFill(
    start_color="FFF2CC",
    end_color="FFF2CC",
    fill_type="solid"
)

RED_FILL = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

# =========================
# 共通関数
# =========================

def safe_copy(src, dst_folder):

    os.makedirs(dst_folder, exist_ok=True)

    filename = os.path.basename(src)

    dst = os.path.join(dst_folder, filename)

    shutil.copy2(src, dst)

    return dst


def safe_copy_unique(src, dst_folder):
    """同名写真がある場合、_001 のように連番を付けて上書きを防ぐ。"""

    os.makedirs(dst_folder, exist_ok=True)

    filename = os.path.basename(src)
    name, ext = os.path.splitext(filename)
    dst = os.path.join(dst_folder, filename)

    count = 1
    while os.path.exists(dst):
        dst = os.path.join(dst_folder, f"{name}_{count:03d}{ext}")
        count += 1

    shutil.copy2(src, dst)

    return dst


def normalize_photo_section(category):
    """AI分類名をデキスパート写真区分へ寄せる。"""

    if category == "出来形測定写真":
        return "出来形管理写真"
    if category in ["作業写真", "重機写真", "黒板写真"]:
        return "施工状況写真"
    if category in ["着工前", "完成写真"]:
        return "着工前及び完成写真"
    return "その他"


def find_tree_row(photo_section, work_type="", type_name="", detail_name=""):
    """master_tree.csv から一番近いツリー行を探す。

    重要:
    着工前及び完成写真のように工種が空の行では、
    細別「着工前」「完成」を優先して探す。
    これをしないと、完成写真が先頭行の「着工前」に入る。
    """

    photo_section = str(photo_section or "").strip()
    work_type = str(work_type or "").strip()
    type_name = str(type_name or "").strip()
    detail_name = str(detail_name or "").strip()

    candidates = [r for r in TREE_MASTER if r.get("写真区分") == photo_section]

    if not candidates:
        return {"写真区分": photo_section, "工種": work_type, "種別": type_name, "細別": detail_name}

    # 1) 細別が明確な場合は最優先。
    #    例: 完成写真 → 着工前及び完成写真 / 完成
    for row in candidates:
        if detail_name and row.get("細別") == detail_name:
            if not work_type or not row.get("工種") or row.get("工種") == work_type:
                return row

    # 2) 完全一致に近い順で探す。
    for row in candidates:
        if work_type and row.get("工種") == work_type and type_name and row.get("種別") == type_name and detail_name and row.get("細別") == detail_name:
            return row

    for row in candidates:
        if work_type and row.get("工種") == work_type and type_name and row.get("種別") == type_name:
            return row

    for row in candidates:
        if work_type and row.get("工種") == work_type:
            return row

    # 3) それでも決まらない場合のみ先頭候補。
    return candidates[0]


def get_tree_folder_from_row(row, fallback_detail=""):
    """写真区分 → 工種 → 種別 → 細別 のフォルダを作る。"""

    section = str(row.get("写真区分", "その他")).strip() or "その他"
    work = str(row.get("工種", "")).strip()
    type_name = str(row.get("種別", "")).strip()
    detail = str(row.get("細別", "")).strip() or str(fallback_detail or "").strip()

    parts = [DEKISPART_TREE, section]

    if work:
        parts.append(work)
    if type_name:
        parts.append(type_name)
    if detail:
        parts.append(detail)

    return os.path.join(*parts)


def copy_to_dekispart_tree(image_path, category, work_type="未分類工種", type_name="", detail_name=""):
    """デキスパートの写真ツリー構造へコピーする。

    master_tree.csv がある場合は、現場ごとの写真区分・工種・種別・細別に合わせる。
    """

    photo_section = normalize_photo_section(category)

    if category == "着工前":
        detail_name = detail_name or "着工前"
    elif category == "完成写真":
        detail_name = detail_name or "完成"
    elif category == "重機写真":
        detail_name = detail_name or "重機"
    elif category == "黒板写真":
        detail_name = detail_name or "黒板"

    row = find_tree_row(photo_section, work_type, type_name, detail_name)

    folder = get_tree_folder_from_row(row, fallback_detail=detail_name)

    return [safe_copy_unique(image_path, folder)]


def image_to_base64(image_path):

    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def extract_json(text):

    text = text.replace("```json", "")
    text = text.replace("```", "")

    match = re.search(r"\{.*\}", text, re.DOTALL)

    if not match:
        return {}

    try:
        return json.loads(match.group())
    except:
        return {}


def to_mm_text(value):

    try:
        return str(int(round(float(value) * 1000)))
    except:
        return ""

# =========================
# AI写真分類
# =========================

def classify_photo(image_path):

    try:
        work_candidates_text = "\n".join([f"- {x}" for x in WORK_TYPE_CANDIDATES])
        tree_hint_text = "\n".join([
            f"- 写真区分:{r.get('写真区分','')} / 工種:{r.get('工種','')} / 種別:{r.get('種別','')} / 細別:{r.get('細別','')}"
            for r in TREE_MASTER
        ])
        base64_image = image_to_base64(image_path)

        response = client.responses.create(
            model="gpt-4.1-mini",
            temperature=0,
            input=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "input_text",
                            "text": f"""
この工事写真を分類してください。

返却はJSONのみ。
説明文は禁止。

分類は必ず次から1つ選んでください。

- 出来形測定写真
- 作業写真
- 重機写真
- 黒板写真
- 完成写真
- 着工前
- 除外

判断基準:
- 出来形測定写真: 測定表、手書き数値、測点番号、田番、出来形管理の写真
- 作業写真: 人が作業している、施工状況が分かる写真
- 重機写真: バックホウ、ローラー、ダンプなど重機が主役
- 黒板写真: 工事黒板が主
- 完成写真: 施工後、完成後の状態
- 着工前: 施工前の状態
- 除外: ピンぼけ、関係ない、判別不能、暗すぎる、地面だけ

追加で、可能なら工種も次から1つ選んでください。
分からない場合は "未分類工種" にしてください。

工種候補:
{work_candidates_text}

現場ツリー候補:
{tree_hint_text}

可能なら type_name と detail_name も現場ツリー候補から選んでください。
分からない場合は空文字で良いです。

返却形式:
{{
  "category": "出来形測定写真",
  "work_type": "整地工",
  "type_name": "整地工",
  "detail_name": "整地仕上げ",
  "reason": "測定表と手書き数値が写っているため"
}}
"""
                        },
                        {
                            "type": "input_image",
                            "image_url": f"data:image/jpeg;base64,{base64_image}"
                        }
                    ]
                }
            ]
        )

        data = extract_json(response.output_text.strip())

        category = str(data.get("category", "除外")).strip()
        work_type = str(data.get("work_type", "未分類工種")).strip()
        type_name = str(data.get("type_name", "")).strip()
        detail_name = str(data.get("detail_name", "")).strip()
        reason = str(data.get("reason", "")).strip()

        if category not in PHOTO_CATEGORIES:
            category = "除外"

        if not work_type:
            work_type = "未分類工種"

        return category, work_type, type_name, detail_name, reason

    except Exception as e:
        return "除外", "未分類工種", "", "", f"分類エラー: {e}"

# =========================
# SSK OCR
# =========================

def gpt_ocr(image_path):

    base64_image = image_to_base64(image_path)

    response = client.responses.create(
        model="gpt-4.1-mini",
        temperature=0,
        input=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": """
このSSK出来形測定写真から情報を読み取ってください。

重要:
- 田番は上部の「田番」欄の数字
- 大きく丸で書かれた手書き数字は田番にしない
- 番号と実測値を必ず対応させる
- No.1を飛ばさない
- 6と8を間違えやすいので注意する

測定項目は次の候補から選択してください。

- 均平度
- 幅
- 厚さ
- 基準高
- 法長
- 延長
- 深さ
- 天端高
- 出来高
- 掘削深
- 盛土高

読む項目:
- 田番
- 測定項目
- 平均値
- 番号と手書き実測値

条件:
- 空欄は無視
- 実測値は小数3桁
- JSONのみ返す
- 説明不要
- 推測しない

返却形式:
{
  "taban": "9",
  "measurement_item": "均平度",
  "average": "1.452",
  "values": [
    {"no": 1, "value": 1.446},
    {"no": 2, "value": 1.456}
  ]
}
"""
                    },
                    {
                        "type": "input_image",
                        "image_url": f"data:image/jpeg;base64,{base64_image}"
                    }
                ]
            }
        ]
    )

    raw_text = response.output_text.strip()

    return raw_text, extract_json(raw_text)

# =========================
# 自動検査
# =========================

def validate_result(taban, measurement_item, values):

    status = "OK"
    reasons = []

    if taban not in expected_counts:
        return "ERROR", ["存在しない田番"]

    if measurement_item not in allowed_items:
        status = "ERROR"
        reasons.append("測定項目不明")

    expected = expected_counts[taban]

    nos = []

    for item in values:
        try:
            nos.append(int(item["no"]))
        except:
            pass

    nos = sorted(nos)

    if len(nos) != expected:
        status = "WARNING"
        reasons.append(f"測点数不一致 {len(nos)}/{expected}")

    expected_nos = list(range(1, expected + 1))

    if nos != expected_nos:
        status = "ERROR"
        reasons.append("測点連番異常")

    for item in values:
        try:
            value = float(item["value"])

            if value < 1.300 or value > 1.600:
                status = "ERROR"
                reasons.append(f"範囲外値 {value}")

        except:
            status = "ERROR"
            reasons.append("数値変換エラー")

    return status, reasons

# =========================
# Excelレポート
# =========================

def create_report_excel(rows):

    wb = Workbook()
    ws = wb.active
    ws.title = "OCRチェック"

    headers = [
        "画像名",
        "分類",
        "田番",
        "測定項目",
        "状態",
        "理由",
        "保存先"
    ]

    ws.append(headers)

    for row in rows:
        ws.append(row)

        status = row[4]

        current_row = ws.max_row

        if status == "OK":
            fill = GREEN_FILL
        elif status == "WARNING":
            fill = YELLOW_FILL
        else:
            fill = RED_FILL

        for cell in ws[current_row]:
            cell.fill = fill

    save_path = os.path.join(REPORTS, "OCRチェック.xlsx")

    wb.save(save_path)

    return save_path

# =========================
# CSV保存
# =========================

def save_taban_csv(save_folder, taban, rows):

    os.makedirs(save_folder, exist_ok=True)

    csv_path = os.path.join(save_folder, f"田番{taban}.csv")

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)

        writer.writerow([
            "測点",
            "工種名",
            "測定項目",
            "設計値",
            "実測値1"
        ])

        writer.writerows(rows)

    return csv_path

# =========================
# 使用写真コピー
# =========================

def copy_selected_by_category(image_path, category, options):

    copied_paths = []

    if category == "出来形測定写真" and options["copy_dekigata"]:
        copied_paths.append(
            safe_copy_unique(image_path, SELECTED_DEKIGATA)
        )

    if category == "作業写真" and options["copy_work"]:
        copied_paths.append(
            safe_copy_unique(image_path, SELECTED_WORK)
        )

    if category == "重機写真" and options["copy_machine"]:
        copied_paths.append(
            safe_copy_unique(image_path, SELECTED_MACHINE)
        )

    if category == "黒板写真" and options["copy_board"]:
        copied_paths.append(
            safe_copy_unique(image_path, SELECTED_BOARD)
        )

    if category == "完成写真" and options["copy_complete"]:
        copied_paths.append(
            safe_copy_unique(image_path, SELECTED_COMPLETE)
        )

    if category == "着工前" and options["copy_before"]:
        copied_paths.append(
            safe_copy_unique(image_path, SELECTED_BEFORE)
        )

    if options["make_submit"]:
        if category in [
            "出来形測定写真",
            "作業写真",
            "重機写真",
            "黒板写真",
            "完成写真",
            "着工前"
        ]:
            copied_paths.append(
                safe_copy_unique(image_path, SELECTED_SUBMIT)
            )

    return copied_paths

# =========================
# メイン処理
# =========================

def run_all(log, options):

    make_folders()

    report_rows = []

    input_folder = options.get("input_folder", IMAGES_ALL)

    if not input_folder or not os.path.isdir(input_folder):
        log(f"読み込み写真フォルダが存在しません: {input_folder}")
        log("ERROR: 読み込み写真フォルダが存在しません")
        return

    image_files = sorted([
        f for f in os.listdir(input_folder)
        if f.lower().endswith((".jpg", ".jpeg", ".png"))
    ])

    if not image_files:
        log(f"画像がありません: {input_folder}")
        log("WARNING: 選択したフォルダに画像がありません")
        return

    log("AI写真分類 + SSK OCR 開始")
    log(f"読み込み写真フォルダ: {input_folder}")
    log(f"対象枚数: {len(image_files)}")

    for index, image_name in enumerate(image_files, start=1):

        image_path = os.path.join(input_folder, image_name)

        log(f"\n[{index}/{len(image_files)}] 処理中: {image_name}")

        category, ai_work_type, ai_type_name, ai_detail_name, reason = classify_photo(image_path)

        log(f"分類: {category}")
        log(f"AI工種: {ai_work_type}")
        if ai_type_name or ai_detail_name:
            log(f"AI種別/細別: {ai_type_name} / {ai_detail_name}")
        log(f"理由: {reason}")

        sorted_path = safe_copy(
            image_path,
            os.path.join(PHOTO_SORTED, category)
        )

        selected_paths = copy_selected_by_category(
            image_path,
            category,
            options
        )

        for path in selected_paths:
            log(f"使用写真コピー: {path}")

        # 出来形測定写真以外は、AIが推定した工種でデキスパート写真ツリーへコピー
        if options.get("make_dekispart_tree") and category != "出来形測定写真":
            tree_paths = copy_to_dekispart_tree(image_path, category, ai_work_type, ai_type_name, ai_detail_name)
            for path in tree_paths:
                log(f"デキスパート写真ツリーコピー: {path}")

        if category != "出来形測定写真":
            report_rows.append([
                image_name,
                category,
                "",
                "",
                "OK",
                reason,
                sorted_path
            ])
            continue

        if not options["ocr_dekigata"]:
            report_rows.append([
                image_name,
                category,
                "",
                "",
                "OK",
                "OCR未実行",
                sorted_path
            ])
            continue

        log("OCR開始")

        try:
            raw_text, result = gpt_ocr(image_path)

            log("GPT OCR結果")
            log(raw_text)

        except Exception as e:
            log(f"OCRエラー: {e}")

            safe_copy(image_path, ERROR_FOLDER)

            report_rows.append([
                image_name,
                category,
                "",
                "",
                "ERROR",
                f"OCRエラー: {e}",
                sorted_path
            ])

            continue

        taban = str(result.get("taban", "")).strip()
        measurement_item = str(result.get("measurement_item", "不明")).strip()
        values = result.get("values", [])

        status, reasons = validate_result(
            taban,
            measurement_item,
            values
        )

        reason_text = " / ".join(reasons)

        log(f"判定: {status}")
        log(f"理由: {reason_text}")

        if status == "OK":
            csv_folder = CSV_OK
        elif status == "WARNING":
            csv_folder = CSV_WARNING
        else:
            csv_folder = CSV_ERROR

        value_dict = {}

        for item in values:
            try:
                no = int(item["no"])
                value = float(item["value"])

                value_dict[no] = value

            except:
                pass

        expected = expected_counts.get(taban, 0)

        design_value = to_mm_text(result.get("average", ""))

        work_type = work_master.get(
            measurement_item,
            ai_work_type if ai_work_type != "未分類工種" else "整地工"
        )

        if options.get("make_dekispart_tree"):
            tree_paths = copy_to_dekispart_tree(image_path, category, work_type, ai_type_name, ai_detail_name)
            for path in tree_paths:
                log(f"デキスパート写真ツリーコピー: {path}")

        rows = []

        for no in range(1, expected + 1):

            if no in value_dict:
                value_text = to_mm_text(value_dict[no])
            else:
                value_text = ""

            rows.append([
                f"No.{no}",
                f"{work_type}田番{taban}",
                measurement_item,
                design_value,
                value_text
            ])

        if taban:
            csv_path = save_taban_csv(
                csv_folder,
                taban,
                rows
            )

            log(f"CSV保存: {csv_path}")
        else:
            csv_path = ""

        report_rows.append([
            image_name,
            category,
            taban,
            measurement_item,
            status,
            reason_text,
            csv_path
        ])

    excel_path = create_report_excel(report_rows)

    log("\n全部完了")
    log(f"Excelレポート: {excel_path}")

    log("完了: 写真分類・使用写真コピー・OCR・CSV作成が完了しました")

# =========================
# GUI
# =========================

root = tk.Tk()
root.title("杏子 all in one")
root.geometry("980x820")

ocr_dekigata_var = tk.BooleanVar(value=True)
copy_dekigata_var = tk.BooleanVar(value=True)
copy_work_var = tk.BooleanVar(value=True)
copy_machine_var = tk.BooleanVar(value=True)
copy_board_var = tk.BooleanVar(value=True)
copy_complete_var = tk.BooleanVar(value=True)
copy_before_var = tk.BooleanVar(value=True)
make_submit_var = tk.BooleanVar(value=True)
make_dekispart_tree_var = tk.BooleanVar(value=True)

input_folder_var = tk.StringVar(value=IMAGES_ALL)

def select_input_folder():
    folder = filedialog.askdirectory(
        title="読み込み写真フォルダを選択",
        initialdir=input_folder_var.get() if os.path.isdir(input_folder_var.get()) else BASE_FOLDER
    )
    if folder:
        input_folder_var.set(folder)

is_running = False

def write_log(message):
    # Tkinterは別スレッドから直接触るとログが混ざることがあるため、root.afterでGUI側に戻す。
    def _append():
        log_text.insert(tk.END, str(message) + "\n")
        log_text.see(tk.END)
    root.after(0, _append)

def set_running_state(running):
    global is_running
    is_running = running

    state = "disabled" if running else "normal"
    start_button.config(state=state)

    # 処理中にフォルダや設定を変えないようにする。
    widgets = [
        select_folder_button,
        ocr_check,
        copy_dekigata_check,
        copy_work_check,
        copy_machine_check,
        copy_board_check,
        copy_complete_check,
        copy_before_check,
        make_submit_check,
        make_dekispart_tree_check,
    ]
    for widget in widgets:
        widget.config(state=state)

def start():
    global is_running

    if is_running:
        write_log("すでに処理中です。完了するまで待ってください。")
        return

    options = {
        "ocr_dekigata": ocr_dekigata_var.get(),
        "copy_dekigata": copy_dekigata_var.get(),
        "copy_work": copy_work_var.get(),
        "copy_machine": copy_machine_var.get(),
        "copy_board": copy_board_var.get(),
        "copy_complete": copy_complete_var.get(),
        "copy_before": copy_before_var.get(),
        "make_submit": make_submit_var.get(),
        "make_dekispart_tree": make_dekispart_tree_var.get(),
        "input_folder": input_folder_var.get()
    }

    set_running_state(True)

    def worker():
        try:
            run_all(write_log, options)
        except Exception as e:
            write_log(f"実行エラー: {e}")
        finally:
            root.after(0, lambda: set_running_state(False))

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()

tk.Label(
    root,
    text="杏子 all in one\n写真分類 → 使用写真コピー → 出来形OCR → デキスパートCSV",
    font=("Meiryo", 14)
).pack(pady=10)

folder_frame = tk.LabelFrame(
    root,
    text="読み込み写真フォルダ",
    padx=10,
    pady=8
)
folder_frame.pack(fill="x", padx=10, pady=5)

select_folder_button = tk.Button(
    folder_frame,
    text="写真フォルダを選択",
    command=select_input_folder,
    width=22
)
select_folder_button.pack(side="left", padx=5)

tk.Label(
    folder_frame,
    textvariable=input_folder_var,
    anchor="w",
    justify="left",
    wraplength=720
).pack(side="left", padx=5, fill="x", expand=True)

frame = tk.LabelFrame(
    root,
    text="実行オプション",
    padx=10,
    pady=10
)

frame.pack(pady=10)

ocr_check = tk.Checkbutton(
    frame,
    text="出来形測定写真をOCRする",
    variable=ocr_dekigata_var
)
ocr_check.grid(row=0, column=0, sticky="w")

copy_dekigata_check = tk.Checkbutton(
    frame,
    text="出来形測定写真を使用写真へコピー",
    variable=copy_dekigata_var
)
copy_dekigata_check.grid(row=1, column=0, sticky="w")

copy_work_check = tk.Checkbutton(
    frame,
    text="作業写真を使用写真へコピー",
    variable=copy_work_var
)
copy_work_check.grid(row=2, column=0, sticky="w")

copy_machine_check = tk.Checkbutton(
    frame,
    text="重機写真を使用写真へコピー",
    variable=copy_machine_var
)
copy_machine_check.grid(row=3, column=0, sticky="w")

copy_board_check = tk.Checkbutton(
    frame,
    text="黒板写真を使用写真へコピー",
    variable=copy_board_var
)
copy_board_check.grid(row=4, column=0, sticky="w")

copy_complete_check = tk.Checkbutton(
    frame,
    text="完成写真を使用写真へコピー",
    variable=copy_complete_var
)
copy_complete_check.grid(row=5, column=0, sticky="w")

copy_before_check = tk.Checkbutton(
    frame,
    text="着工前写真を使用写真へコピー",
    variable=copy_before_var
)
copy_before_check.grid(row=6, column=0, sticky="w")

make_submit_check = tk.Checkbutton(
    frame,
    text="提出候補フォルダにもコピー",
    variable=make_submit_var
)
make_submit_check.grid(row=7, column=0, sticky="w")

make_dekispart_tree_check = tk.Checkbutton(
    frame,
    text="デキスパート写真ツリー用フォルダも作成",
    variable=make_dekispart_tree_var
)
make_dekispart_tree_check.grid(row=8, column=0, sticky="w")

start_button = tk.Button(
    root,
    text="AI写真分類 + OCR + 使用写真抽出 開始",
    command=start,
    bg="blue",
    fg="white",
    height=2,
    width=40
)
start_button.pack(pady=10)

log_text = scrolledtext.ScrolledText(
    root,
    width=130,
    height=35
)

log_text.pack(padx=10, pady=10)

root.mainloop()





