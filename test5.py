from PIL import Image
import pytesseract
import re
import pandas as pd
import os
from openpyxl import load_workbook

print("OCR → デキスパートExcel 自動入力開始")

# Tesseractパス
pytesseract.pytesseract.tesseract_cmd = (
    r"C:\Program Files\Tesseract-OCR\tesseract.exe"
)

# 画像フォルダ
image_folder = "images"

# Excel
input_file = "sample.xlsx"
output_file = "完成_自動入力.xlsx"

# 抽出値保存
values = []

# 画像処理
for filename in os.listdir(image_folder):

    if filename.lower().endswith((".png", ".jpg", ".jpeg")):

        image_file = os.path.join(image_folder, filename)

        print("処理中:", image_file)

        # 画像読込
        img = Image.open(image_file)

        # グレースケール化
        img = img.convert("L")

        # OCR確認画像保存
        debug_file = "確認_" + filename
        img.save(debug_file)

        print("確認画像保存:", debug_file)

        # OCR
        text = pytesseract.image_to_string(
            img,
            lang="eng",
            config="--psm 6 -c tessedit_char_whitelist=0123456789."
        )

        print("OCR結果")
        print(text)

        # 数値抽出
        raw_values = re.findall(
            r"[0-9]?\.[0-9]{1,3}|[0-9]{4}",
            text
        )

        print("抽出数値")
        print(raw_values)

        image_values = []

        for v in raw_values:

            # .403 → 1.403
            if v.startswith("."):
                v = "1" + v

            # 1405 → 1.405
            elif "." not in v and len(v) == 4:
                v = str(float(v) / 1000)

            # 小数桁補正
            if "." in v:

                left, right = v.split(".")

                # 1.41 → 1.410
                if len(right) == 2:
                    right = right + "0"

                # 1.4 → 1.400
                elif len(right) == 1:
                    right = right + "00"

                v = left + "." + right

            # 数値化
            num = float(v)

            # 範囲フィルタ
            if 1.3 <= num <= 1.5:

                values.append(num)
                image_values.append(num)

        print(filename, "→", image_values)

print("復元値")
print(values)

# =========================
# Excel入力
# =========================

wb = load_workbook(input_file)

ws = wb["測定結果表"]

start_row = 10
target_col = "C"

for i, value in enumerate(values):

    row = start_row + i * 2

    cell = ws[f"{target_col}{row}"]

    cell.value = value

    # 1.400 表示固定
    cell.number_format = "0.000"

    print(f"{target_col}{row} = {value}")

# 保存
wb.save(output_file)

print("保存しました:", output_file)

# =========================
# CSV出力
# =========================

csv_data = []

for i, value in enumerate(values):

    csv_data.append([
        f"GE-{i+1}",
        f"{value:.3f}"
    ])

df = pd.DataFrame(csv_data)

df.to_csv(
    "dekispart.csv",
    index=False,
    header=False,
    encoding="cp932"
)

print("CSV保存完了: dekispart.csv")

